from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl 
from openpyxl import Workbook
import pathlib

# hacemos la raiz sobre la que vamos a trabajar#
raiz  = Tk()
raiz.title("Proyecto") #le damos un nombre a la barra de cerrado#
raiz.geometry('1000x600+400+200') #definimos el tamaño de la pestaña#
raiz.resizable(False,False) #con este comando no permitimos modifical el tamaño de la pestaña#
raiz.configure(bg = "#757575") #damos color al fondo#
#añadimos el icono#
icono_ = PhotoImage(file="Icono2.png")
raiz.iconphoto(False,icono_)

file = pathlib.Path('DATA.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet= file.active
    sheet['A1']= "NOMBRE DE LA MICRO EMPRESA"
    sheet['B1']= "NOMBRE DEL PROPIETARIO"
    sheet['C1']= "CORREO ELECTRONICO"
    sheet['D1']= "UBICACION"
    sheet['E1']= "CAMPO DE ACCION"
    sheet['F1'] = "TIPO DE VENTA"
    sheet['A4'] = "NOMBRE DEL PRODUCTO"
    sheet['B4'] = "CANTIDAD"
    sheet['C4'] = "DESCRIPCION"
    sheet['D4'] = "PRECIO"

    file.save('DATA.xlsx')
def clear():
    nombremic.set('')
    nombrepro.set('')
    correoel.set('')
    ubicacion.set('')
    nombrepoduc.set('')
    #cantidad# set 0
    Descripcion.delete(1.0,END)

def agregarDatos():
    nombre= nombremic.get()
    nombrep=nombrepro.get()
    correo=correoel.get()
    ubi=ubicacion.get()
    campoA=campo_caja.get()
    tipo=tipo_caja.get()

    file=openpyxl.load_workbook('DATA.xlsx')
    sheet= file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=nombre)
    sheet.cell(column=2, row=sheet.max_row, value=nombrep)
    sheet.cell(column=3, row=sheet.max_row, value=correo)
    sheet.cell(column=4, row=sheet.max_row, value=ubi)
    sheet.cell(column=5, row=sheet.max_row, value=campoA)
    sheet.cell(column=6 , row=sheet.max_row, value=tipo)
    file.save('DATA.xlsx')
    messagebox.showinfo('INFORMACION','DATOS AÑADIDOS')

    nombremic.set('')
    nombrepro.set('')
    correoel.set('')
    ubicacion.set('')
    nombrepoduc.set('')

def agregarProducto():

    nombreprd = nombrepoduc.get()
    cantida = cantidad.get()
    descrip = Descripcion.get(1.0,END)
    prec = precio.get()
    file = openpyxl.load_workbook('DATA.xlsx')
    sheet = file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=nombreprd)
    sheet.cell(column=2, row=sheet.max_row, value=cantida)
    sheet.cell(column=3, row=sheet.max_row, value=descrip)
    sheet.cell(column=4, row=sheet.max_row, value=prec)
    file.save('DATA.xlsx')
    nombrepoduc.set('')
    cantidad.set(0)
    Descripcion.delete(1.0, END)
    precio.set(0)
    messagebox.showinfo('INFORMACION', 'PRODUCTOS AÑADIDOS')
# Titulo #
Label(raiz,text="Af App",font = "cursive 19",bg ="#757575", fg="#000000").place(x=460,y=20)
# division #
Label(raiz,text='-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------', font= "arial 6", bg= "#757575", fg= "#A6A6A6").place(x=10, y=50)
# nombre de la micro empresa #
Label(raiz,text="Nombre de la micro empresa :",font = "alegreya 11",bg ="#757575", fg="#000000").place(x=70,y=80)
# nombre del propietario 120#
Label(raiz,text='Nombre del propietario :', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=550, y=80)
# ingrese un correo que este activo 160 #
Label(raiz,text='Correo electronico :', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=70, y=120)
# campo de accion al que esta enfocado la empresa 200#
Label(raiz,text='Campo de acción :', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=70, y=160)
# Ubicacion // solo dentro de colombia #
Label(raiz,text='Ubicación :', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=550, y=120)
Label(raiz,text='únicamente dentro de la nación Colombiana', font= "alegreya 7", bg= "#757575", fg= "#000000").place(x=550, y=140)
#tipo de venta, informal o formal#
Label(raiz,text='Tipo de venta :', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=550, y=160)
# division #
Label(raiz,text='-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------', font= "arial 6", bg= "#757575", fg= "#A6A6A6").place(x=10, y=200)
Label(raiz,text='-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------', font= "arial 6", bg= "#757575", fg= "#A6A6A6").place(x=10, y=215)
# nombre del producto #
Label(raiz,text='Nombre de producto :', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=70, y=240)
# cantidad a ingresar en inventario #
Label(raiz,text='Cantidad :', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=550, y=240)
# descripcion del producto #
Label(raiz,text='Descripción:', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=70, y=280)
# boton agregar #
Label(raiz,text='Precio:', font= "alegreya 11", bg= "#757575", fg= "#000000").place(x=550, y=280)
#division #
Label(raiz,text='-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------', font= "arial 6", bg= "#757575", fg= "#A6A6A6").place(x=10, y=340)
Label(raiz,text='-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------', font= "arial 6", bg= "#757575", fg= "#A6A6A6").place(x=10, y=355)
#cuadros de texto#
nombremic= StringVar()
nombrepro = StringVar()
correoel= StringVar()
ubicacion = StringVar()
nombrepoduc = StringVar()
cantidad = IntVar()
descripcion = StringVar()
precio = IntVar()
nomMingreso = Entry(raiz, textvariable=nombremic,width=25 ,bd=1)
nomMingreso.place(x= 300,y=80)
nomPropieta = Entry(raiz, textvariable=nombrepro,width=25 ,bd=1)
nomPropieta.place(x= 735,y=80)
CorreoE = Entry(raiz, textvariable=correoel,width=25 ,bd=1)
CorreoE.place(x= 220,y=120)
Ubicacion = Entry(raiz, textvariable=ubicacion,width=25 ,bd=1)
Ubicacion.place(x= 638,y=120)
Nombreproduc = Entry(raiz, textvariable=nombrepoduc,width=25 ,bd=1)
Nombreproduc.place(x= 235,y=240)
Cantidad = Entry(raiz, textvariable=cantidad,width=5 ,bd=1)
Cantidad.place(x= 638,y=240)
Descripcion = Text(raiz,width=30,height=3,bd=1)
Descripcion.place(x=165,y=280)
Precio = Entry(raiz,textvariable=precio,width=10,bd=1)
Precio.place(x=605,y=280)
#Casillas con opciones
campo_caja = Combobox(raiz, values=['Supermercado','Puesto de comida','Vestuario','Tegnogia','Panaderia','Otros'],font="alegreya 11",state='r',width=15)
campo_caja.place(x=210,y=160)
tipo_caja = Combobox(raiz,values=['Formal','Informal'],font="alegreya 11",state='r',width=15)
tipo_caja.place(x=665,y=160)
#Boton#
Button(raiz,text="Agregar Datos",bg= "#B3B3B3",width=15,height=2,command= agregarDatos).place(x=700, y=400)
Button(raiz,text= "Borrar todo", bg="#B3B3B3",width=15,height=2,command=clear).place(x=100,y=400)
Button(raiz,text= "Salir", bg="#B3B3B3",width=3,height=1,command=lambda:raiz.destroy()).place(x=940,y=15)
Button(raiz,text= "Agregar Productos", bg="#B3B3B3",width=15,height=2,command=agregarProducto).place(x=400,y=400)
Label(raiz,text='-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------', font= "arial 6", bg= "#757575", fg= "#A6A6A6").place(x=10, y=460)
Label(raiz,text='-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------', font= "arial 6", bg= "#757575", fg= "#A6A6A6").place(x=10, y=475)
raiz.mainloop()
